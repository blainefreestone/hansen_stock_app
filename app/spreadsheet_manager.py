import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import subprocess
import os
from typing import Dict, Any
from datetime import datetime

from app.data_processor import DataProcessor
from app.formatting import FormattingRuleFactory, FormatStyle
from app.stock_data_fetcher import StockDataFetcher

color_map = {
    "red": "FFFF0000",
    "green": "FF00FF00",
    "yellow": "FFFFFF00",
    "black": "FF000000",
    "lightblue": "FFADD8E6",
}

class SpreadSheetManager:
    def __init__(self, api_key: str):
        self.data_processor = DataProcessor()
        self.stock_data_fetcher = StockDataFetcher(api_key)

    def create_excel_file(self, user_input: Dict[str, Any]) -> str:
        # Fetch stock data
        stock_data = self.stock_data_fetcher.fetch_daily_stock_data(
            user_input["symbol"],
            user_input["start_date"],
            user_input["end_date"]
        )
        
        if stock_data is None:
            raise ValueError("Failed to fetch stock data")

        # Process data
        percent_changes = self.data_processor.calculate_daily_percent_changes(stock_data)
        
        # Add percent changes to stock data
        for date, change in percent_changes.items():
            stock_data[date]['percent_change'] = change

        # Remove 'stock_symbol' key from stock_data
        stock_data_filtered = {k: v for k, v in stock_data.items() if k != 'stock_symbol'}

        # Create DataFrame
        df = pd.DataFrame.from_dict(stock_data_filtered, orient='index')
        df = df.sort_index()

        # Apply formatting rules
        consecutive_rule = FormattingRuleFactory().consecutive_change_rule(
            user_input["consecutive_change"]["days"],
            user_input["consecutive_change"]["direction"],
            "close",
            FormatStyle("close", "yellow", bold=True)
        )
        
        threshold_rule = FormattingRuleFactory().threshold_change_rule(
            user_input["daily_threshold"]["percent"],
            user_input["daily_threshold"]["direction"],
            "percent_change",
            FormatStyle("percent_change", "red" if user_input["daily_threshold"]["direction"] == "higher" else "green", bold=True)
        )
        
        cumulative_rule = FormattingRuleFactory().cumulative_change_rule(
            user_input["period_change"]["days"],
            user_input["period_change"]["percent"],
            ["open", "high", "low"],
            FormatStyle(["open", "high", "low"], "lightblue", bold=True)
        )

        # Create Excel file
        file_name = f"{user_input['symbol']}_stock_data.xlsx"
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock Data')
            workbook = writer.book
            worksheet = writer.sheets['Stock Data']

            # Apply formatting
            for rule in [consecutive_rule, threshold_rule, cumulative_rule]:
                formatting = rule.apply(stock_data_filtered)
                for date, style in formatting.items():
                    if style:
                        row = df.index.get_loc(date) + 2  # +2 because Excel is 1-indexed and we have a header row
                        cols = [df.columns.get_loc(col) + 2 for col in style.columns] if isinstance(style.columns, list) else [df.columns.get_loc(style.columns) + 2]
                        for col in cols:
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = PatternFill(start_color=color_map[style.background_color], end_color=color_map[style.background_color], fill_type="solid")
                            cell.font = Font(color=color_map[style.font_color], bold=style.bold)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max(max_length + 2, 12)  # Ensure minimum width of 12
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        return file_name

    def open_excel_file(self, file_name: str):
        if os.name == 'nt':  # For Windows
            os.startfile(file_name)
        elif os.name == 'posix':  # For macOS and Linux
            opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
            subprocess.call([opener, file_name])

    def process_and_open(self, user_input: Dict[str, Any]):
        # try:
        #     file_name = self.create_excel_file(user_input)
        #     self.open_excel_file(file_name)
        # except Exception as e:
        #     print(f"An error occurred: {str(e)}")
        #     # You might want to show this error in the GUI instead of printing
        file_name = self.create_excel_file(user_input)
        self.open_excel_file(file_name)